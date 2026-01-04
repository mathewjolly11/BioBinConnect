import csv

from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

def generate_pdf_report(response, title, headers, data):
    """
    Generate a PDF report using ReportLab
    response: HttpResponse object
    title: Report title (string)
    headers: List of column headers
    data: List of lists containing row data
    """
    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1 # Center
    )
    
    # Title
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table Data
    table_data = [headers] + data
    
    # Table table
    # Calculate column widths (simple dynamic sizing)
    col_widths = [None] * len(headers) # Auto
    
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')), # Header Green
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(t)
    doc.build(elements)

def generate_excel_report(response, title, headers, data):
    """
    Generate an Excel report using OpenPyXL
    response: HttpResponse object
    title: Report title (string)
    headers: List of column headers
    data: List of lists containing row data
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Title
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True, color="10B981")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Date
    ws['A2'] = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    # Data
    for row_num, row_data in enumerate(data, 5):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = str(cell_value) # Convert to string to avoid potential formatting issues
            
    # Auto-adjust column widths
    for col_num, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_num)].width = length + 2
        
    wb.save(response)
